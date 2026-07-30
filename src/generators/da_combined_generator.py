"""Combined Data Architecture generator.

Produces a single multi-sheet Excel workbook (data-architecture.xlsx) with all
DA-01 through DA-07 sheets, plus 3 PPTX diagram files.
"""

from pathlib import Path

from src.models.project import ProjectInfo, DataEntity, DataRelationship, EnumDefinition
from src.generators.excel_generator import create_workbook, add_sheet, save_workbook
from src.analyzers.data_entity_analyzer import analyze_data_entities
from src.analyzers.enum_analyzer import analyze_enums
from src.analyzers.crud_analyzer import analyze_crud
from src.utils.naming import generate_encoding
from src.i18n import t, get_headers


def generate_data_architecture(project: ProjectInfo, output_dir: Path, locale: str = "zh") -> Path:
    """Generate complete data architecture workbook with all DA-01~DA-07 sheets.

    Runs all data architecture analyzers and produces a single Excel file
    with multiple sheets.

    Returns path to the generated Excel file.
    """
    # Step 1: Run analyzers
    entities, relationships = analyze_data_entities(project)
    enums = analyze_enums(project)
    crud_records = analyze_crud(project, entities)

    # Step 2: Create workbook
    wb = create_workbook()

    # Step 3: Generate each sheet
    _add_guide_sheet(wb, project, locale)
    _add_da01_sheet(wb, project, entities, locale)
    _add_da02_sheet(wb, project, entities, locale)
    _add_da03_sheet(wb, project, entities, locale)
    _add_da04_sheet(wb, project, entities, locale)
    _add_da05_sheet(wb, project, entities, crud_records, locale)
    _add_da06_sheet(wb, project, entities, crud_records, locale)
    _add_da07_sheet(wb, project, enums, locale)
    _add_cross_references(wb, locale)

    # Step 4: Save
    filename = t("file.da_combined", locale)
    return save_workbook(wb, output_dir / filename)


def generate_all_da(project: ProjectInfo, output_dir: Path, locale: str = "zh") -> list[Path]:
    """Generate data architecture artifacts: combined Excel + 3 PPTX diagrams.

    Returns list of all generated file paths.
    """
    from src.generators.da_cdm_generator import generate_da_cdm
    from src.generators.da_ldm_generator import generate_da_ldm
    from src.generators.da_flow_generator import generate_da_flow

    # Run analyzers once
    entities, relationships = analyze_data_entities(project)
    enums = analyze_enums(project)
    crud_records = analyze_crud(project, entities)

    outputs: list[Path] = []

    # Combined multi-sheet Excel workbook
    combined_path = _generate_combined_workbook(project, entities, enums, crud_records, output_dir, locale)
    outputs.append(combined_path)

    # PPTX diagram files
    outputs.append(generate_da_cdm(project, entities, relationships, output_dir, locale))
    outputs.append(generate_da_ldm(project, entities, relationships, output_dir, locale))
    outputs.append(generate_da_flow(project, entities, crud_records, output_dir, locale))

    return outputs


def _generate_combined_workbook(
    project: ProjectInfo,
    entities: list[DataEntity],
    enums: list[EnumDefinition],
    crud_records: list[dict],
    output_dir: Path,
    locale: str = "zh",
) -> Path:
    """Generate the combined multi-sheet workbook using pre-computed analysis results."""
    wb = create_workbook()

    _add_guide_sheet(wb, project, locale)
    _add_da01_sheet(wb, project, entities, locale)
    _add_da02_sheet(wb, project, entities, locale)
    _add_da03_sheet(wb, project, entities, locale)
    _add_da04_sheet(wb, project, entities, locale)
    _add_da05_sheet(wb, project, entities, crud_records, locale)
    _add_da06_sheet(wb, project, entities, crud_records, locale)
    _add_da07_sheet(wb, project, enums, locale)
    _add_cross_references(wb, locale)

    filename = t("file.da_combined", locale)
    return save_workbook(wb, output_dir / filename)


# ---------------------------------------------------------------------------
# Sheet helpers — replicate row-building logic from individual generators
# ---------------------------------------------------------------------------


def _get_translation_keys() -> set[str]:
    """Return the set of known translation keys for checking existence."""
    from src.i18n import TRANSLATIONS
    return set(TRANSLATIONS.keys())


def _add_guide_sheet(wb, project: ProjectInfo, locale: str) -> None:
    """Add a guide/description sheet as the first sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date

    sheet_name = t("sheet.da_guide", locale)
    ws = wb.create_sheet(title=sheet_name, index=0)
    font_name = t("font.name", locale)

    # Title
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value=t("da.guide_title", locale))
    title_cell.font = Font(name=font_name, size=14, bold=True)

    # Project info
    ws.cell(row=3, column=1, value=t("da.guide_project", locale)).font = Font(name=font_name, size=10, bold=True)
    ws.cell(row=3, column=2, value=project.name).font = Font(name=font_name, size=10)
    ws.cell(row=4, column=1, value=t("da.guide_generated", locale)).font = Font(name=font_name, size=10, bold=True)
    ws.cell(row=4, column=2, value=date.today().isoformat()).font = Font(name=font_name, size=10)

    # Sheet description table
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    cell_font = Font(name=font_name, size=9)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        t("da.guide_sheet_col", locale),
        t("da.guide_desc_col", locale),
        t("da.guide_ref_col", locale),
    ]
    start_row = 6
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Sheet descriptions and cross-references
    sheet_info = [
        (t("sheet.da01", locale), t("da.guide_da01_desc", locale), "DA-02, DA-05"),
        (t("sheet.da02", locale), t("da.guide_da02_desc", locale), "DA-01, DA-03, DA-07"),
        (t("sheet.da03", locale), t("da.guide_da03_desc", locale), "DA-02, DA-04"),
        (t("sheet.da04", locale), t("da.guide_da04_desc", locale), "DA-03, DA-06"),
        (t("sheet.da05", locale), t("da.guide_da05_desc", locale), "DA-01, DA-02"),
        (t("sheet.da06", locale), t("da.guide_da06_desc", locale), "DA-04"),
        (t("sheet.da07", locale), t("da.guide_da07_desc", locale), "DA-02"),
    ]

    for row_idx, (name, desc, refs) in enumerate(sheet_info, start_row + 1):
        ws.cell(row=row_idx, column=1, value=name).font = cell_font
        ws.cell(row=row_idx, column=2, value=desc).font = cell_font
        ws.cell(row=row_idx, column=3, value=refs).font = cell_font
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22


def _add_cross_references(wb, locale: str) -> None:
    """Add hyperlinks in each DA sheet's header row linking to related sheets."""
    sheet_relations = {
        t("sheet.da01", locale): [t("sheet.da02", locale), t("sheet.da05", locale)],
        t("sheet.da02", locale): [t("sheet.da01", locale), t("sheet.da03", locale), t("sheet.da07", locale)],
        t("sheet.da03", locale): [t("sheet.da02", locale), t("sheet.da04", locale)],
        t("sheet.da04", locale): [t("sheet.da03", locale), t("sheet.da06", locale)],
        t("sheet.da05", locale): [t("sheet.da01", locale), t("sheet.da02", locale)],
        t("sheet.da06", locale): [t("sheet.da04", locale)],
        t("sheet.da07", locale): [t("sheet.da02", locale)],
    }

    for sheet_name, related_sheets in sheet_relations.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Find first empty column after headers
        max_col = ws.max_column
        ref_col = max_col + 2

        from openpyxl.styles import Font
        font_name = t("font.name", locale)
        link_font = Font(name=font_name, size=8, color="0563C1", underline="single")

        for idx, ref_sheet in enumerate(related_sheets):
            if ref_sheet in wb.sheetnames:
                cell = ws.cell(row=1, column=ref_col + idx)
                cell.value = ref_sheet
                cell.hyperlink = f"#'{ref_sheet}'!A1"
                cell.font = link_font


def _add_da01_sheet(wb, project: ProjectInfo, entities: list[DataEntity], locale: str) -> None:
    """DA-01 Conceptual Entity List."""
    headers = get_headers("da01", locale)
    rows = []

    # Group entities by data_domain
    domain_order: list[str] = []
    domain_entities: dict[str, list[DataEntity]] = {}
    for entity in entities:
        domain = entity.data_domain or entity.module_name
        if domain not in domain_entities:
            domain_order.append(domain)
            domain_entities[domain] = []
        domain_entities[domain].append(entity)

    entity_idx = 0
    for domain_idx, domain in enumerate(domain_order, 1):
        domain_id = generate_encoding("DD", domain_idx)
        domain_name = domain

        for entity in domain_entities[domain]:
            entity_idx += 1
            entity_id = generate_encoding("DE", entity_idx)

            # Determine data category
            if entity.data_category:
                data_category = (
                    t(f"val.{entity.data_category}", locale)
                    if f"val.{entity.data_category}" in _get_translation_keys()
                    else entity.data_category
                )
            else:
                data_category = t("val.transaction_data", locale)

            rows.append([
                domain_id,
                domain_name,
                entity_id,
                entity.class_name,
                entity.class_name,  # business object = class_name for now
                t("val.yes", locale),
                data_category,
                entity.module_name,
            ])

    add_sheet(wb, t("sheet.da01", locale), headers, rows, locale=locale)


def _add_da02_sheet(wb, project: ProjectInfo, entities: list[DataEntity], locale: str) -> None:
    """DA-02 Logical Entity List."""
    headers = get_headers("da02", locale)
    rows = []

    for entity_idx, entity in enumerate(entities, 1):
        data_domain = entity.data_domain or entity.module_name
        concept_entity_id = generate_encoding("DE", entity_idx)
        concept_entity_name = entity.class_name
        logical_entity_id = generate_encoding("DL", entity_idx)
        logical_entity_name = entity.class_name

        for field in entity.fields:
            attr_name = field.name
            attr_code = field.column_name if field.column_name else field.name
            data_type = field.java_type
            is_pk = t("val.yes", locale) if field.is_primary_key else t("val.no", locale)
            is_fk = t("val.yes", locale) if field.is_foreign_key else t("val.no", locale)
            is_not_null = t("val.yes", locale) if field.is_primary_key or not field.is_nullable else t("val.no", locale)

            rows.append([
                data_domain,
                concept_entity_id,
                concept_entity_name,
                logical_entity_id,
                logical_entity_name,
                attr_name,
                attr_code,
                data_type,
                is_pk,
                is_fk,
                is_not_null,
            ])

    add_sheet(wb, t("sheet.da02", locale), headers, rows, locale=locale)


def _add_da03_sheet(wb, project: ProjectInfo, entities: list[DataEntity], locale: str) -> None:
    """DA-03 Physical Entity List."""
    headers = get_headers("da03", locale)
    rows = []

    for entity_idx, entity in enumerate(entities, 1):
        data_domain = entity.data_domain or entity.module_name
        logical_entity_id = generate_encoding("DL", entity_idx)
        logical_entity_name = entity.class_name
        physical_entity_id = generate_encoding("DP", entity_idx)
        physical_entity_name = entity.class_name

        for field in entity.fields:
            field_name = field.name
            field_code = field.column_name if field.column_name else field.name
            data_type = field.java_type

            rows.append([
                data_domain,
                logical_entity_id,
                logical_entity_name,
                physical_entity_id,
                physical_entity_name,
                field_name,
                field_code,
                data_type,
            ])

    add_sheet(wb, t("sheet.da03", locale), headers, rows, locale=locale)


def _add_da04_sheet(wb, project: ProjectInfo, entities: list[DataEntity], locale: str) -> None:
    """DA-04 Database Table List."""
    headers = get_headers("da04", locale)
    rows = []

    for entity_idx, entity in enumerate(entities, 1):
        data_domain = entity.data_domain or entity.module_name
        physical_entity_id = generate_encoding("DP", entity_idx)
        physical_entity_name = entity.class_name
        table_id = generate_encoding("DT", entity_idx)
        table_name = entity.table_name

        for field in entity.fields:
            field_name = field.name
            field_code = field.column_name if field.column_name else field.name

            rows.append([
                data_domain,
                physical_entity_id,
                physical_entity_name,
                table_id,
                table_name,
                field_name,
                field_code,
                project.name,
                "",  # db_type - unknown from scan
            ])

    add_sheet(wb, t("sheet.da04", locale), headers, rows, locale=locale)


def _add_da05_sheet(wb, project: ProjectInfo, entities: list[DataEntity], crud_records: list[dict], locale: str) -> None:
    """DA-05 Data Source List."""
    OPERATION_KEYS = {
        "C": "val.create",
        "R": "val.read",
        "U": "val.update",
        "D": "val.delete",
    }

    headers = get_headers("da05", locale)
    rows = []

    # Build entity index lookup: class_name -> 1-based index
    entity_index: dict[str, int] = {}
    for idx, entity in enumerate(entities, 1):
        entity_index[entity.class_name] = idx

    for record in crud_records:
        entity_name = record.get("entity", "")
        entity_idx = entity_index.get(entity_name, 0)

        # Generate encoding IDs
        concept_entity_id = generate_encoding("DE", entity_idx) if entity_idx else ""
        logical_entity_id = generate_encoding("DL", entity_idx) if entity_idx else ""

        # Translate operation type
        operation = record.get("operation", "")
        operation_key = OPERATION_KEYS.get(operation, "")
        operation_display = t(operation_key, locale) if operation_key else operation

        rows.append([
            record.get("data_domain", ""),
            concept_entity_id,
            entity_name,
            logical_entity_id,
            entity_name,
            operation_display,
            record.get("app_name", ""),
            record.get("module", ""),
            record.get("function", ""),
        ])

    add_sheet(wb, t("sheet.da05", locale), headers, rows, locale=locale)


def _add_da06_sheet(wb, project: ProjectInfo, entities: list[DataEntity], crud_records: list[dict], locale: str) -> None:
    """DA-06 Table-Function Relationship."""
    OPERATION_KEYS = {
        "C": "val.create",
        "R": "val.read",
        "U": "val.update",
        "D": "val.delete",
    }

    headers = get_headers("da06", locale)
    rows = []

    # Build entity lookup: class_name -> (table_name, 1-based index)
    entity_lookup: dict[str, tuple[str, int]] = {}
    for idx, entity in enumerate(entities, 1):
        entity_lookup[entity.class_name] = (entity.table_name, idx)

    for record in crud_records:
        entity_name = record.get("entity", "")
        lookup = entity_lookup.get(entity_name)

        # Skip records where entity is not found in the entities list
        if lookup is None:
            continue

        table_name, table_idx = lookup

        # Generate table encoding
        table_id = generate_encoding("DT", table_idx)

        # Translate operation type
        operation = record.get("operation", "")
        operation_key = OPERATION_KEYS.get(operation, "")
        operation_display = t(operation_key, locale) if operation_key else operation

        rows.append([
            project.name,
            table_id,
            table_name,
            operation_display,
            record.get("app_name", ""),
            record.get("module", ""),
            record.get("function", ""),
        ])

    add_sheet(wb, t("sheet.da06", locale), headers, rows, locale=locale)


def _add_da07_sheet(wb, project: ProjectInfo, enums: list[EnumDefinition], locale: str) -> None:
    """DA-07 Data Dictionary."""
    headers = get_headers("da07", locale)
    rows = []

    for enum_idx, enum_def in enumerate(enums, 1):
        enum_type_id = generate_encoding("DD", enum_idx)

        for value in enum_def.values:
            enum_cn_name = value.get("label") or value.get("name", "")
            enum_value = value.get("value") or value.get("name", "")
            enum_en_name = value.get("name", "")

            rows.append([
                enum_type_id,
                enum_def.class_name,
                enum_cn_name,
                enum_value,
                enum_en_name,
                t("val.enabled", locale),
                "",
            ])

    add_sheet(wb, t("sheet.da07", locale), headers, rows, locale=locale)
