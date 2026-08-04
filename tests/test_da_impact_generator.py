"""Tests for DA-IMPACT impact analysis Excel report generator."""

import pytest
from pathlib import Path
from openpyxl import load_workbook

from src.analyzers.lineage_graph import LineageGraph
from src.models.project import FieldLineage
from src.generators.da_impact_generator import generate_da_impact, _sanitize_field_name


@pytest.fixture
def sample_graph():
    """Build a simple lineage graph for testing."""
    graph = LineageGraph()
    lineages = [
        FieldLineage(
            source_system="order-service", source_entity="Order", source_field="id",
            target_system="payment-service", target_entity="Payment", target_field="orderId",
            mapping_type="setter", transform_expr="", trigger_mode="realtime",
            confidence=0.9, evidence="Payment.setOrderId(order.getId())",
        ),
        FieldLineage(
            source_system="payment-service", source_entity="Payment", source_field="orderId",
            target_system="report-service", target_entity="Report", target_field="paymentOrderId",
            mapping_type="BeanUtils", transform_expr="", trigger_mode="realtime",
            confidence=0.7, evidence="BeanUtils.copyProperties(payment, report)",
        ),
        FieldLineage(
            source_system="order-service", source_entity="Order", source_field="id",
            target_system="audit-service", target_entity="AuditLog", target_field="entityId",
            mapping_type="MapStruct", transform_expr="", trigger_mode="realtime",
            confidence=1.0, evidence="@Mapping(source=\"id\", target=\"entityId\")",
        ),
    ]
    graph.build_from_lineages(lineages)
    return graph


def test_generate_da_impact_basic(sample_graph, tmp_path):
    """Test basic DA-IMPACT Excel generation."""
    result_path = generate_da_impact(
        sample_graph, "order-service.Order.id", tmp_path, locale="zh"
    )

    assert result_path.exists()
    assert "DA-IMPACT" in result_path.name
    assert result_path.suffix == ".xlsx"

    wb = load_workbook(result_path)
    ws = wb.active
    assert ws.title == "DA-IMPACT 影响分析报告"

    # Header row + data rows: Order.id has 3 paths
    # path 1: order-service.Order.id -> payment-service.Payment.orderId
    # path 2: order-service.Order.id -> payment-service.Payment.orderId -> report-service.Report.paymentOrderId
    # path 3: order-service.Order.id -> audit-service.AuditLog.entityId
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 3

    # Check that all paths are present
    paths_col = [row[4] for row in data_rows]
    assert any("payment-service.Payment.orderId" in p and "report-service.Report.paymentOrderId" in p for p in paths_col)
    assert any("audit-service.AuditLog.entityId" in p for p in paths_col)


def test_generate_da_impact_empty_result(sample_graph, tmp_path):
    """Test DA-IMPACT with a leaf node (no downstream)."""
    result_path = generate_da_impact(
        sample_graph, "report-service.Report.paymentOrderId", tmp_path, locale="zh"
    )

    assert result_path.exists()
    wb = load_workbook(result_path)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 0


def test_generate_da_impact_nonexistent_field(sample_graph, tmp_path):
    """Test DA-IMPACT with a field not in the graph."""
    result_path = generate_da_impact(
        sample_graph, "unknown.Entity.field", tmp_path, locale="zh"
    )

    assert result_path.exists()
    wb = load_workbook(result_path)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 0


def test_generate_da_impact_english_locale(sample_graph, tmp_path):
    """Test DA-IMPACT with English locale."""
    result_path = generate_da_impact(
        sample_graph, "order-service.Order.id", tmp_path, locale="en"
    )

    assert result_path.exists()
    assert "Impact_Analysis" in result_path.name

    wb = load_workbook(result_path)
    ws = wb.active
    assert ws.title == "DA-IMPACT Impact Analysis"

    # Verify English headers
    headers = [cell.value for cell in ws[1]]
    assert "Affected Field" in headers
    assert "System" in headers
    assert "Propagation Path" in headers


def test_generate_da_impact_confidence_and_depth(sample_graph, tmp_path):
    """Test that depth and min confidence are computed correctly."""
    result_path = generate_da_impact(
        sample_graph, "order-service.Order.id", tmp_path, locale="zh"
    )

    wb = load_workbook(result_path)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Find the 2-hop path (depth=2): Order.id -> Payment.orderId -> Report.paymentOrderId
    two_hop = [r for r in data_rows if r[5] == 2]
    assert len(two_hop) == 1
    # Min confidence along this path: min(0.9, 0.7) = 0.7
    assert two_hop[0][7] == 0.7

    # The direct MapStruct path (depth=1) has confidence 1.0
    mapstruct_row = [r for r in data_rows if r[0] == "audit-service.AuditLog.entityId" and r[5] == 1]
    assert len(mapstruct_row) == 1
    assert mapstruct_row[0][7] == 1.0


def test_sanitize_field_name():
    """Test filename sanitization."""
    assert _sanitize_field_name("order-service.Order.id") == "order-service_Order_id"
    assert _sanitize_field_name("a" * 100) == "a" * 50
    assert _sanitize_field_name("sys.entity.field<>") == "sys_entity_field__"
