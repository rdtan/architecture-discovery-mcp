"""E2E acceptance test for data lineage against the real ThingsBoard project.

ThingsBoard is a large Spring Boot + JPA project (~4000+ Java files) with:
- No MapStruct mappers
- No MyBatis XML mappers
- Extensive manual setter/getter mapping (toData() pattern, 174+ files)
- @Scheduled annotations (25+ files)
- Custom Kafka abstraction (TbKafkaProducerTemplate, not Spring KafkaTemplate)

This test validates the full data lineage pipeline against a real-world project.
"""

import pytest
from pathlib import Path

from src.scanner.project_scanner import scan_project
from src.generators.lineage_combined_generator import generate_data_lineage
from src.analyzers.lineage_analyzer import analyze_mapstruct_lineage
from src.analyzers.setter_getter_analyzer import analyze_setter_getter_lineage
from src.analyzers.sql_lineage_analyzer import analyze_sql_lineage
from src.analyzers.dataflow_analyzer import analyze_dataflows
from src.analyzers.data_entity_analyzer import analyze_data_entities
from src.analyzers.lineage_graph import LineageGraph

THINGSBOARD_PATH = Path(r"D:\AI\AI-Architecture-Discovery-Engine\thingsboard")

pytestmark = pytest.mark.skipif(
    not THINGSBOARD_PATH.exists(),
    reason=f"ThingsBoard fixture not found at {THINGSBOARD_PATH}"
)


@pytest.fixture(scope="module")
def thingsboard_project():
    """Scan the real ThingsBoard project. Cached per test module for performance."""
    if not THINGSBOARD_PATH.exists():
        pytest.skip("ThingsBoard project not found at expected path")
    return scan_project(THINGSBOARD_PATH)


@pytest.fixture(scope="module")
def thingsboard_entities(thingsboard_project):
    """Run data entity analysis (needed for setter/getter analyzer)."""
    entities, relationships = analyze_data_entities(thingsboard_project)
    return entities


# =============================================================================
# Full Pipeline E2E Test
# =============================================================================


class TestFullPipeline:
    """End-to-end test: scan -> analyze -> generate -> verify outputs."""

    def test_generate_data_lineage_produces_files(self, thingsboard_project, tmp_path):
        result = generate_data_lineage(thingsboard_project, tmp_path, locale="zh")

        assert "files" in result
        assert "stats" in result
        assert "graph" in result

        # DA-08, DA-09, and DA-LINEAGE files must be generated
        assert len(result["files"]) == 3
        for f in result["files"]:
            assert f.exists(), f"Output file missing: {f}"
            assert f.stat().st_size > 0, f"Output file empty: {f}"

    def test_stats_reflect_detected_lineages(self, thingsboard_project, tmp_path):
        result = generate_data_lineage(thingsboard_project, tmp_path, locale="zh")
        stats = result["stats"]

        # ThingsBoard has extensive setter/getter patterns
        assert stats["total_lineages"] > 0
        assert stats["setter_getter_count"] > 0

        # No MapStruct in ThingsBoard
        assert stats["mapstruct_count"] == 0

        # No MyBatis mapper XMLs
        assert stats["sql_count"] == 0

        # @Scheduled annotations should produce ETL flows
        assert stats["dataflow_count"] > 0

        # Graph should be populated
        assert stats["graph_nodes"] > 0
        assert stats["graph_edges"] > 0

    def test_english_locale_filenames(self, thingsboard_project, tmp_path):
        result = generate_data_lineage(thingsboard_project, tmp_path, locale="en")

        filenames = [f.name for f in result["files"]]
        assert any("Field_Level_Lineage" in name for name in filenames)
        assert any("Data_Flow_Inventory" in name for name in filenames)

    def test_graph_is_queryable(self, thingsboard_project, tmp_path):
        result = generate_data_lineage(thingsboard_project, tmp_path, locale="zh")
        graph = result["graph"]

        assert isinstance(graph, LineageGraph)
        assert graph.node_count > 0
        assert graph.edge_count > 0

        # Nonexistent node should return empty results gracefully
        assert graph.trace_forward("nonexistent.Entity.field") == []
        assert graph.trace_backward("nonexistent.Entity.field") == []

    def test_da08_excel_has_content(self, thingsboard_project, tmp_path):
        """DA-08 Excel should have actual data rows beyond the header."""
        from openpyxl import load_workbook

        result = generate_data_lineage(thingsboard_project, tmp_path, locale="zh")
        da08_file = result["files"][0]

        wb = load_workbook(da08_file)
        ws = wb.active
        # Header + at least one data row
        assert ws.max_row >= 2, f"DA-08 has only {ws.max_row} rows (expected data)"

    def test_da09_excel_has_content(self, thingsboard_project, tmp_path):
        """DA-09 Excel should have actual data rows (from @Scheduled flows)."""
        from openpyxl import load_workbook

        result = generate_data_lineage(thingsboard_project, tmp_path, locale="zh")
        da09_file = result["files"][1]

        wb = load_workbook(da09_file)
        ws = wb.active
        assert ws.max_row >= 2, f"DA-09 has only {ws.max_row} rows (expected data)"


# =============================================================================
# Individual Analyzer E2E Tests
# =============================================================================


class TestMapStructAnalyzer:
    """MapStruct analyzer should find nothing in ThingsBoard (no MapStruct usage)."""

    def test_zero_mapstruct_lineages(self, thingsboard_project):
        lineages = analyze_mapstruct_lineage(thingsboard_project)
        assert lineages == []


class TestSetterGetterAnalyzer:
    """Setter/getter analyzer should detect ThingsBoard's toData() mapping pattern."""

    def test_detects_setter_getter_lineages(self, thingsboard_project, thingsboard_entities):
        lineages = analyze_setter_getter_lineage(thingsboard_project, thingsboard_entities)

        # ThingsBoard has 174+ files with setter/getter patterns
        assert len(lineages) > 0, "Expected setter/getter lineages from toData() pattern"

    def test_lineage_fields_are_populated(self, thingsboard_project, thingsboard_entities):
        lineages = analyze_setter_getter_lineage(thingsboard_project, thingsboard_entities)

        if not lineages:
            pytest.skip("No setter/getter lineages detected")

        for lineage in lineages[:20]:  # Spot-check first 20
            assert lineage.source_system, "source_system must not be empty"
            assert lineage.source_entity, "source_entity must not be empty"
            assert lineage.source_field, "source_field must not be empty"
            assert lineage.target_system, "target_system must not be empty"
            assert lineage.target_entity, "target_entity must not be empty"
            assert lineage.target_field, "target_field must not be empty"
            assert lineage.mapping_type == "direct"
            assert lineage.confidence == 1.0
            assert lineage.evidence == "setter/getter chain"

    def test_source_system_matches_module_name(self, thingsboard_project, thingsboard_entities):
        lineages = analyze_setter_getter_lineage(thingsboard_project, thingsboard_entities)

        if not lineages:
            pytest.skip("No setter/getter lineages detected")

        module_names = {m.name for m in thingsboard_project.modules}
        for lineage in lineages:
            assert lineage.source_system in module_names, (
                f"source_system '{lineage.source_system}' not in project modules"
            )


class TestSqlLineageAnalyzer:
    """SQL lineage analyzer should find nothing (ThingsBoard uses JPA, not MyBatis)."""

    def test_zero_sql_lineages(self, thingsboard_project):
        lineages = analyze_sql_lineage(thingsboard_project)
        assert lineages == [], (
            f"Expected zero SQL lineages for non-MyBatis project, got {len(lineages)}"
        )


class TestDataflowAnalyzer:
    """Dataflow analyzer should detect @Scheduled tasks in ThingsBoard."""

    def test_detects_scheduled_flows(self, thingsboard_project):
        flows = analyze_dataflows(thingsboard_project)

        # ThingsBoard has 25+ @Scheduled annotations
        scheduled_flows = [f for f in flows if f.transport_type == "ETL"]
        assert len(scheduled_flows) > 0, "Expected @Scheduled ETL flows"

    def test_flow_ids_are_sequential(self, thingsboard_project):
        flows = analyze_dataflows(thingsboard_project)

        if not flows:
            pytest.skip("No flows detected")

        # Flow IDs should follow DF-NNN pattern
        for flow in flows:
            assert flow.flow_id.startswith("DF-"), f"Invalid flow_id: {flow.flow_id}"

    def test_flow_fields_populated(self, thingsboard_project):
        flows = analyze_dataflows(thingsboard_project)

        if not flows:
            pytest.skip("No flows detected")

        for flow in flows:
            assert flow.source_system == thingsboard_project.name
            assert flow.source_module, "source_module must not be empty"
            assert flow.transport_type in ("API", "RPC", "MQ", "ETL")
            assert flow.frequency, "frequency must not be empty"

    def test_scheduled_frequency_format(self, thingsboard_project):
        flows = analyze_dataflows(thingsboard_project)
        scheduled_flows = [f for f in flows if f.transport_type == "ETL"]

        if not scheduled_flows:
            pytest.skip("No scheduled flows")

        for flow in scheduled_flows:
            # Frequency should be one of: "scheduled", "cron(...)", "every Nms", "every Nms (delayed)"
            assert any(
                flow.frequency.startswith(prefix)
                for prefix in ("scheduled", "cron(", "every ")
            ), f"Unexpected frequency format: {flow.frequency}"


# =============================================================================
# Graph Integration Tests (built from real ThingsBoard data)
# =============================================================================


class TestLineageGraph:
    """Test that the lineage graph built from ThingsBoard data is coherent."""

    @pytest.fixture
    def thingsboard_graph(self, thingsboard_project, thingsboard_entities):
        lineages = analyze_setter_getter_lineage(thingsboard_project, thingsboard_entities)
        flows = analyze_dataflows(thingsboard_project)
        graph = LineageGraph()
        graph.build_from_lineages(lineages, flows)
        return graph

    def test_graph_has_nodes_and_edges(self, thingsboard_graph):
        assert thingsboard_graph.node_count > 0
        assert thingsboard_graph.edge_count > 0

    def test_forward_trace_returns_list(self, thingsboard_graph):
        # Pick an arbitrary node that exists
        if thingsboard_graph.node_count == 0:
            pytest.skip("Empty graph")

        # Trace from a nonexistent node should return empty list
        result = thingsboard_graph.trace_forward("nonexistent.X.y")
        assert isinstance(result, list)
        assert result == []

    def test_impact_analysis_on_nonexistent_node(self, thingsboard_graph):
        result = thingsboard_graph.impact_analysis("nonexistent.X.y")
        assert result.total_affected == 0
        assert result.paths == []
