import pytest
from pathlib import Path
from src.scanner.project_scanner import scan_project
from src.analyzers.module_analyzer import analyze_modules
from src.analyzers.api_analyzer import analyze_apis
from src.analyzers.integration_analyzer import analyze_integrations
from src.generators.aa01_generator import generate_aa01
from src.generators.aa02_generator import generate_aa02
from src.generators.aa03_generator import generate_aa03
from src.generators.aa04_generator import generate_aa04
from src.generators.aa05_generator import generate_aa05
from src.generators.aa07_generator import generate_aa07
from src.generators.aa08_generator import generate_aa08


def test_full_pipeline(sample_project_path, tmp_path):
    # Scan
    project = scan_project(sample_project_path)
    assert project.name == "ecommerce-platform"
    assert len(project.modules) == 2

    # Analyze
    project = analyze_modules(project)
    endpoints = analyze_apis(project)
    integrations = analyze_integrations(project)

    assert len(endpoints) >= 5
    assert len(integrations) >= 1

    # Generate all artifacts
    artifacts = [
        generate_aa01(project, tmp_path),
        generate_aa02(project, endpoints, tmp_path),
        generate_aa03(project, endpoints, tmp_path),
        generate_aa04(project, endpoints, tmp_path),
        generate_aa05(project, integrations, tmp_path),
        generate_aa07(project, tmp_path),
        generate_aa08(project, integrations, tmp_path),
    ]

    # Verify all outputs exist and are non-empty
    for artifact in artifacts:
        assert artifact.exists(), f"Missing: {artifact}"
        assert artifact.stat().st_size > 0, f"Empty: {artifact}"

    # Verify Excel content
    from openpyxl import load_workbook
    wb = load_workbook(artifacts[0])  # AA-01
    ws = wb.active
    assert ws.max_row >= 3  # header + data

    # Verify correct artifact types
    excel_count = sum(1 for a in artifacts if a.suffix == ".xlsx")
    pptx_count = sum(1 for a in artifacts if a.suffix == ".pptx")
    assert excel_count == 5
    assert pptx_count == 2
