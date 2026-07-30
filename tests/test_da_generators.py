import pytest
from pathlib import Path
from src.models.project import ProjectInfo, DataEntity, EntityField, EnumDefinition, Module, ApiEndpoint
from src.generators.da01_generator import generate_da01
from src.generators.da02_generator import generate_da02
from src.generators.da03_generator import generate_da03
from src.generators.da04_generator import generate_da04
from src.generators.da05_generator import generate_da05
from src.generators.da06_generator import generate_da06
from src.generators.da07_generator import generate_da07
from src.analyzers.crud_analyzer import analyze_crud


def test_generate_da01(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="userId", java_type="Long", column_name="user_id"),
            ]
        ),
        DataEntity(
            module_name="order-service",
            class_name="OrderItem",
            table_name="order_items",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
            ]
        ),
    ]
    output = generate_da01(project, entities, tmp_path)
    assert output.exists()
    assert "DA-01" in output.name

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # 2 data rows (one per entity)
    assert ws.max_row == 3  # 1 header + 2 data rows


def test_generate_da01_multiple_domains(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-domain",
            fields=[EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True)],
        ),
        DataEntity(
            module_name="user-service",
            class_name="User",
            table_name="users",
            data_domain="user-domain",
            fields=[EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True)],
        ),
    ]
    output = generate_da01(project, entities, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Check domain IDs are different
    assert ws.cell(row=2, column=1).value == "DD-001"
    assert ws.cell(row=3, column=1).value == "DD-002"
    # Check entity IDs are sequential
    assert ws.cell(row=2, column=3).value == "DE-001"
    assert ws.cell(row=3, column=3).value == "DE-002"


def test_generate_da01_uses_module_name_when_no_domain(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="payment-service",
            class_name="Payment",
            table_name="payments",
            data_domain="",
            fields=[EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True)],
        ),
    ]
    output = generate_da01(project, entities, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Domain name should fall back to module_name
    assert ws.cell(row=2, column=2).value == "payment-service"


def test_generate_da01_en_locale(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True)],
        ),
    ]
    output = generate_da01(project, entities, tmp_path, locale="en")
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Check English headers
    assert ws.cell(row=1, column=1).value == "Data Domain ID"
    assert ws.cell(row=1, column=6).value == "Is Core Entity"
    # Check English values
    assert ws.cell(row=2, column=6).value == "Yes"
    assert ws.cell(row=2, column=7).value == "Transaction Data"


# --- DA-07 Tests ---


def test_generate_da07(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    enums = [
        EnumDefinition(
            module_name="order-service",
            class_name="OrderStatus",
            values=[
                {"name": "PENDING", "value": "pending", "label": "待处理"},
                {"name": "COMPLETED", "value": "completed", "label": "已完成"},
                {"name": "CANCELLED", "value": "cancelled", "label": "已取消"},
            ]
        ),
    ]
    output = generate_da07(project, enums, tmp_path)
    assert output.exists()
    assert "DA-07" in output.name

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # 3 data rows (one per enum value)
    assert ws.max_row == 4  # 1 header + 3 values
    # Check first data row
    assert ws.cell(2, 2).value == "OrderStatus"  # enum type name
    assert ws.cell(2, 4).value == "pending"  # enum value


def test_generate_da07_multiple_enums(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    enums = [
        EnumDefinition(
            module_name="order-service",
            class_name="OrderStatus",
            values=[
                {"name": "PENDING", "value": "0", "label": "待处理"},
                {"name": "DONE", "value": "1", "label": "完成"},
            ]
        ),
        EnumDefinition(
            module_name="user-service",
            class_name="UserRole",
            values=[
                {"name": "ADMIN", "value": "admin", "label": "管理员"},
            ]
        ),
    ]
    output = generate_da07(project, enums, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row == 4  # 1 header + 3 values total

    # First enum gets DD-001
    assert ws.cell(2, 1).value == "DD-001"
    assert ws.cell(3, 1).value == "DD-001"
    # Second enum gets DD-002
    assert ws.cell(4, 1).value == "DD-002"
    assert ws.cell(4, 2).value == "UserRole"
    assert ws.cell(4, 5).value == "ADMIN"


def test_generate_da07_fallback_no_label(tmp_path):
    """When label is empty, fall back to name for enum_cn_name."""
    project = ProjectInfo(name="test-project", path=Path("."))
    enums = [
        EnumDefinition(
            module_name="service",
            class_name="Direction",
            values=[
                {"name": "UP", "value": "up", "label": ""},
                {"name": "DOWN", "value": "", "label": ""},
            ]
        ),
    ]
    output = generate_da07(project, enums, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # label empty -> fallback to name
    assert ws.cell(2, 3).value == "UP"
    # value empty -> fallback to name
    assert ws.cell(3, 4).value == "DOWN"


def test_generate_da07_english_locale(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    enums = [
        EnumDefinition(
            module_name="service",
            class_name="Status",
            values=[
                {"name": "ACTIVE", "value": "1", "label": "活跃"},
            ]
        ),
    ]
    output = generate_da07(project, enums, tmp_path, locale="en")
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Status column should use English
    assert ws.cell(2, 6).value == "Enabled"


# --- DA-02 Tests ---


def test_generate_da02(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="userId", java_type="Long", column_name="user_id"),
                EntityField(name="totalAmount", java_type="Double", column_name="total_amount"),
            ]
        ),
    ]
    output = generate_da02(project, entities, tmp_path)
    assert output.exists()
    assert "DA-02" in output.name

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # 3 data rows (one per field)
    assert ws.max_row == 4  # 1 header + 3 data rows


def test_generate_da02_multiple_entities(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="userId", java_type="Long", column_name="user_id"),
            ]
        ),
        DataEntity(
            module_name="user-service",
            class_name="User",
            table_name="users",
            data_domain="user-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="name", java_type="String", column_name="name", is_nullable=False),
            ]
        ),
    ]
    output = generate_da02(project, entities, tmp_path)
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # 4 data rows (2 fields per entity x 2 entities)
    assert ws.max_row == 5  # 1 header + 4 data rows

    # Check encoding values
    assert ws.cell(row=2, column=2).value == "DE-001"  # First entity concept ID
    assert ws.cell(row=2, column=4).value == "DL-001"  # First entity logical ID
    assert ws.cell(row=4, column=2).value == "DE-002"  # Second entity concept ID
    assert ws.cell(row=4, column=4).value == "DL-002"  # Second entity logical ID


def test_generate_da02_pk_fk_nullable(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="userId", java_type="Long", column_name="user_id", is_foreign_key=True, is_nullable=False),
                EntityField(name="note", java_type="String", column_name="note", is_nullable=True),
            ]
        ),
    ]
    output = generate_da02(project, entities, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active

    # Row 2: id - PK=yes, FK=no, NotNull=yes (PK implies not null)
    assert ws.cell(row=2, column=9).value == "是"   # is_pk
    assert ws.cell(row=2, column=10).value == "否"  # is_fk
    assert ws.cell(row=2, column=11).value == "是"  # is_not_null

    # Row 3: userId - PK=no, FK=yes, NotNull=yes (not nullable)
    assert ws.cell(row=3, column=9).value == "否"   # is_pk
    assert ws.cell(row=3, column=10).value == "是"  # is_fk
    assert ws.cell(row=3, column=11).value == "是"  # is_not_null

    # Row 4: note - PK=no, FK=no, NotNull=no (nullable)
    assert ws.cell(row=4, column=9).value == "否"   # is_pk
    assert ws.cell(row=4, column=10).value == "否"  # is_fk
    assert ws.cell(row=4, column=11).value == "否"  # is_not_null


# --- DA-03 Tests ---


def test_generate_da03(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="userId", java_type="Long", column_name="user_id"),
            ]
        ),
    ]
    output = generate_da03(project, entities, tmp_path)
    assert output.exists()
    assert "DA-03" in output.name
    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row == 3  # 1 header + 2 field rows


def test_generate_da03_encoding(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
            ]
        ),
        DataEntity(
            module_name="user-service",
            class_name="User",
            table_name="users",
            data_domain="user-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
            ]
        ),
    ]
    output = generate_da03(project, entities, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Check logical and physical entity encodings
    assert ws.cell(row=2, column=2).value == "DL-001"
    assert ws.cell(row=2, column=4).value == "DP-001"
    assert ws.cell(row=3, column=2).value == "DL-002"
    assert ws.cell(row=3, column=4).value == "DP-002"


def test_generate_da03_field_values(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-domain",
            fields=[
                EntityField(name="userId", java_type="Long", column_name="user_id"),
            ]
        ),
    ]
    output = generate_da03(project, entities, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Check field-level values
    assert ws.cell(row=2, column=1).value == "order-domain"  # data_domain
    assert ws.cell(row=2, column=3).value == "Order"  # logical entity name
    assert ws.cell(row=2, column=5).value == "Order"  # physical entity name
    assert ws.cell(row=2, column=6).value == "userId"  # field name
    assert ws.cell(row=2, column=7).value == "user_id"  # field code
    assert ws.cell(row=2, column=8).value == "Long"  # data type


# --- DA-04 Tests ---


def test_generate_da04(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="userId", java_type="Long", column_name="user_id"),
            ]
        ),
    ]
    output = generate_da04(project, entities, tmp_path)
    assert output.exists()
    assert "DA-04" in output.name
    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row == 3  # 1 header + 2 field rows
    # Check table name is from @Table annotation
    assert ws.cell(2, 5).value == "orders"


def test_generate_da04_system_name(tmp_path):
    project = ProjectInfo(name="my-system", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
            ]
        ),
    ]
    output = generate_da04(project, entities, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # system_name should be project.name
    assert ws.cell(row=2, column=8).value == "my-system"
    # db_type should be empty (openpyxl returns None for empty cells)
    assert ws.cell(row=2, column=9).value in ("", None)


def test_generate_da04_encoding(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
            ]
        ),
        DataEntity(
            module_name="user-service",
            class_name="User",
            table_name="users",
            data_domain="user-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
            ]
        ),
    ]
    output = generate_da04(project, entities, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Check physical entity and table encodings
    assert ws.cell(row=2, column=2).value == "DP-001"
    assert ws.cell(row=2, column=4).value == "DT-001"
    assert ws.cell(row=3, column=2).value == "DP-002"
    assert ws.cell(row=3, column=4).value == "DT-002"


# --- CRUD Analyzer Tests ---


def test_analyze_crud_with_api_endpoints():
    """Test CRUD analysis using API endpoints."""
    project = ProjectInfo(name="test-project", path=Path("."))
    project.api_endpoints = [
        ApiEndpoint(
            module_name="order-service",
            class_name="OrderController",
            method_name="createOrder",
            http_method="POST",
            path="/api/orders",
        ),
        ApiEndpoint(
            module_name="order-service",
            class_name="OrderController",
            method_name="getOrder",
            http_method="GET",
            path="/api/orders/{id}",
        ),
        ApiEndpoint(
            module_name="order-service",
            class_name="OrderController",
            method_name="updateStatus",
            http_method="PUT",
            path="/api/orders/{id}/status",
        ),
        ApiEndpoint(
            module_name="order-service",
            class_name="OrderController",
            method_name="deleteOrder",
            http_method="DELETE",
            path="/api/orders/{id}",
        ),
    ]
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[],
        ),
    ]

    crud = analyze_crud(project, entities)
    assert len(crud) == 4

    operations = {rec["operation"] for rec in crud}
    assert operations == {"C", "R", "U", "D"}

    for rec in crud:
        assert rec["entity"] == "Order"
        assert rec["data_domain"] == "order-service"
        assert rec["app_name"] == "order-service"
        assert rec["module"] == "OrderController"
        assert "operation" in rec
        assert rec["operation"] in ("C", "R", "U", "D")


def test_analyze_crud_fallback_no_endpoints():
    """Test CRUD analysis falls back to module names when no endpoints."""
    project = ProjectInfo(name="test-project", path=Path("."))
    project.modules = [
        Module(
            name="order-service",
            path=Path("."),
            controllers=["OrderController"],
            repositories=["OrderRepository"],
        ),
    ]
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[],
        ),
    ]

    crud = analyze_crud(project, entities)
    assert len(crud) > 0
    for rec in crud:
        assert rec["entity"] == "Order"
        assert rec["operation"] in ("C", "R", "U", "D")


def test_analyze_crud_empty_entities():
    """Test CRUD analysis with no entities returns empty."""
    project = ProjectInfo(name="test-project", path=Path("."))
    crud = analyze_crud(project, [])
    assert crud == []


def test_analyze_crud_no_match():
    """Test CRUD analysis when controller names don't match any entity."""
    project = ProjectInfo(name="test-project", path=Path("."))
    project.api_endpoints = [
        ApiEndpoint(
            module_name="misc-service",
            class_name="HealthController",
            method_name="healthCheck",
            http_method="GET",
            path="/health",
        ),
    ]
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[],
        ),
    ]

    crud = analyze_crud(project, entities)
    # HealthController doesn't match Order entity
    assert len(crud) == 0


def test_analyze_crud_with_fixture_project(tmp_path):
    """Test CRUD analysis with the fixture sample project."""
    sample_path = Path("tests/fixtures/sample-java-project")
    if not sample_path.exists():
        pytest.skip("Sample project fixture not available")

    from src.scanner.project_scanner import scan_project
    from src.analyzers.module_analyzer import analyze_modules
    from src.analyzers.api_analyzer import analyze_apis
    from src.analyzers.data_entity_analyzer import analyze_data_entities

    project = scan_project(sample_path)
    project = analyze_modules(project)
    project.api_endpoints = analyze_apis(project)
    entities, _ = analyze_data_entities(project)

    crud = analyze_crud(project, entities)
    assert len(crud) > 0
    for rec in crud:
        assert "entity" in rec
        assert "operation" in rec
        assert rec["operation"] in ("C", "R", "U", "D")


# --- DA-05 Tests ---


def test_generate_da05(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[],
        ),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "R", "app_name": "order-service", "module": "OrderController", "function": "getOrder"},
    ]
    output = generate_da05(project, entities, crud_records, tmp_path)
    assert output.exists()
    assert "DA-05" in output.name

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row == 3  # 1 header + 2 CRUD rows


def test_generate_da05_encoding(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
        DataEntity(module_name="user-service", class_name="User", table_name="users", data_domain="user-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
        {"entity": "User", "data_domain": "user-service", "operation": "R", "app_name": "user-service", "module": "UserController", "function": "getUser"},
    ]
    output = generate_da05(project, entities, crud_records, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Row 2: Order -> DE-001, DL-001
    assert ws.cell(row=2, column=2).value == "DE-001"
    assert ws.cell(row=2, column=4).value == "DL-001"
    assert ws.cell(row=2, column=3).value == "Order"
    assert ws.cell(row=2, column=5).value == "Order"
    # Row 3: User -> DE-002, DL-002
    assert ws.cell(row=3, column=2).value == "DE-002"
    assert ws.cell(row=3, column=4).value == "DL-002"


def test_generate_da05_operation_translation(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "R", "app_name": "order-service", "module": "OrderController", "function": "getOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "U", "app_name": "order-service", "module": "OrderController", "function": "updateOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "D", "app_name": "order-service", "module": "OrderController", "function": "deleteOrder"},
    ]

    # Test Chinese locale
    output = generate_da05(project, entities, crud_records, tmp_path)
    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.cell(row=2, column=6).value == "创建"
    assert ws.cell(row=3, column=6).value == "读取"
    assert ws.cell(row=4, column=6).value == "修改"
    assert ws.cell(row=5, column=6).value == "删除"


def test_generate_da05_en_locale(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
    ]
    output = generate_da05(project, entities, crud_records, tmp_path, locale="en")
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Check English header
    assert ws.cell(row=1, column=6).value == "Operation Type"
    # Check English operation value
    assert ws.cell(row=2, column=6).value == "Create"


def test_generate_da05_empty_records(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = []
    crud_records = []
    output = generate_da05(project, entities, crud_records, tmp_path)
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row == 1  # header only


# --- DA-06 Tests ---


def test_generate_da06(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "R", "app_name": "order-service", "module": "OrderController", "function": "getOrder"},
    ]
    output = generate_da06(project, entities, crud_records, tmp_path)
    assert output.exists()
    assert "DA-06" in output.name
    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row == 3  # 1 header + 2 rows
    # Check table name comes from @Table annotation
    assert ws.cell(2, 3).value == "orders"
    # Check source system
    assert ws.cell(2, 1).value == "test-project"


def test_generate_da06_skips_unknown_entities(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
        {"entity": "Unknown", "data_domain": "other", "operation": "R", "app_name": "other-service", "module": "OtherController", "function": "getData"},
    ]
    output = generate_da06(project, entities, crud_records, tmp_path)
    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Only 1 data row (Unknown entity skipped)
    assert ws.max_row == 2


def test_generate_da06_encoding(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
        DataEntity(module_name="user-service", class_name="User", table_name="users", data_domain="user-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
        {"entity": "User", "data_domain": "user-service", "operation": "R", "app_name": "user-service", "module": "UserController", "function": "getUser"},
    ]
    output = generate_da06(project, entities, crud_records, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Row 2: Order -> DT-001
    assert ws.cell(row=2, column=2).value == "DT-001"
    assert ws.cell(row=2, column=3).value == "orders"
    # Row 3: User -> DT-002
    assert ws.cell(row=3, column=2).value == "DT-002"
    assert ws.cell(row=3, column=3).value == "users"


def test_generate_da06_operation_translation(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "R", "app_name": "order-service", "module": "OrderController", "function": "getOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "U", "app_name": "order-service", "module": "OrderController", "function": "updateOrder"},
        {"entity": "Order", "data_domain": "order-service", "operation": "D", "app_name": "order-service", "module": "OrderController", "function": "deleteOrder"},
    ]

    # Test Chinese locale
    output = generate_da06(project, entities, crud_records, tmp_path)
    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.cell(row=2, column=4).value == "创建"
    assert ws.cell(row=3, column=4).value == "读取"
    assert ws.cell(row=4, column=4).value == "修改"
    assert ws.cell(row=5, column=4).value == "删除"


def test_generate_da06_en_locale(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(module_name="order-service", class_name="Order", table_name="orders", data_domain="order-service", fields=[]),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
    ]
    output = generate_da06(project, entities, crud_records, tmp_path, locale="en")
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Check English header
    assert ws.cell(row=1, column=4).value == "Operation Type"
    # Check English operation value
    assert ws.cell(row=2, column=4).value == "Create"


def test_generate_da06_empty_records(tmp_path):
    project = ProjectInfo(name="test-project", path=Path("."))
    entities = []
    crud_records = []
    output = generate_da06(project, entities, crud_records, tmp_path)
    assert output.exists()

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row == 1  # header only


# --- Combined DA Generator Tests ---


def test_generate_data_architecture_combined(tmp_path):
    """Test combined multi-sheet Excel generation."""
    from src.generators.da_combined_generator import generate_data_architecture

    sample_path = Path(__file__).parent / "fixtures" / "sample-java-project"
    if not sample_path.exists():
        pytest.skip("Sample project fixture not available")

    from src.scanner.project_scanner import scan_project
    from src.analyzers.module_analyzer import analyze_modules

    project = scan_project(sample_path)
    project = analyze_modules(project)

    output = generate_data_architecture(project, tmp_path)
    assert output.exists()
    assert output.name == "data-architecture.xlsx"

    import openpyxl
    wb = openpyxl.load_workbook(output)
    # Should have 8 sheets (Guide + DA-01 through DA-07)
    assert len(wb.sheetnames) == 8


def test_generate_data_architecture_combined_empty(tmp_path):
    """Test combined generation with empty project produces 7 sheets with headers only."""
    from src.generators.da_combined_generator import _generate_combined_workbook

    project = ProjectInfo(name="empty-project", path=Path("."))
    entities = []
    enums = []
    crud_records = []

    output = _generate_combined_workbook(project, entities, enums, crud_records, tmp_path)
    assert output.exists()
    assert output.name == "data-architecture.xlsx"

    import openpyxl
    wb = openpyxl.load_workbook(output)
    assert len(wb.sheetnames) == 8


def test_generate_data_architecture_combined_with_data(tmp_path):
    """Test combined generation produces correct sheet data."""
    from src.generators.da_combined_generator import _generate_combined_workbook

    project = ProjectInfo(name="test-project", path=Path("."))
    entities = [
        DataEntity(
            module_name="order-service",
            class_name="Order",
            table_name="orders",
            data_domain="order-service",
            fields=[
                EntityField(name="id", java_type="Long", column_name="id", is_primary_key=True),
                EntityField(name="userId", java_type="Long", column_name="user_id"),
            ]
        ),
    ]
    enums = [
        EnumDefinition(
            module_name="order-service",
            class_name="OrderStatus",
            values=[
                {"name": "PENDING", "value": "0", "label": "待处理"},
                {"name": "DONE", "value": "1", "label": "完成"},
            ]
        ),
    ]
    crud_records = [
        {"entity": "Order", "data_domain": "order-service", "operation": "C", "app_name": "order-service", "module": "OrderController", "function": "createOrder"},
    ]

    output = _generate_combined_workbook(project, entities, enums, crud_records, tmp_path)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    assert len(wb.sheetnames) == 8

    # DA-01 sheet should have 1 data row (one entity)
    ws_da01 = wb.worksheets[1]
    assert ws_da01.max_row == 2  # 1 header + 1 entity

    # DA-02 sheet should have 2 data rows (2 fields)
    ws_da02 = wb.worksheets[2]
    assert ws_da02.max_row == 3  # 1 header + 2 fields

    # DA-07 sheet should have 2 data rows (2 enum values)
    ws_da07 = wb.worksheets[7]
    assert ws_da07.max_row == 3  # 1 header + 2 enum values


def test_generate_all_da(tmp_path):
    """Test full DA pipeline generates all individual + combined files."""
    from src.generators.da_combined_generator import generate_all_da

    sample_path = Path(__file__).parent / "fixtures" / "sample-java-project"
    if not sample_path.exists():
        pytest.skip("Sample project fixture not available")

    from src.scanner.project_scanner import scan_project
    from src.analyzers.module_analyzer import analyze_modules

    project = scan_project(sample_path)
    project = analyze_modules(project)

    outputs = generate_all_da(project, tmp_path)
    # 7 individual Excel + 3 PPTX diagrams + 1 combined Excel = 11 files
    assert len(outputs) == 4
    for path in outputs:
        assert path.exists()
