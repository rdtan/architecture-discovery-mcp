import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock
from src.utils.naming import camel_to_words, generate_encoding
from src.generators.excel_generator import create_workbook, add_sheet, save_workbook
from src.scanner.project_scanner import scan_project
from src.analyzers.module_analyzer import analyze_modules
from src.analyzers.api_analyzer import analyze_apis
from src.analyzers.integration_analyzer import analyze_integrations
from src.generators.aa01_generator import generate_aa01
from src.generators.aa02_generator import generate_aa02
from src.generators.aa03_generator import generate_aa03
from src.generators.aa04_generator import generate_aa04
from src.generators.aa05_generator import generate_aa05
from src.generators.aa07_generator import generate_aa07
from src.generators.aa08_generator import generate_aa08
from src.llm.enhancer import LLMEnhancer
from src.models.project import Module, Integration, IntegrationType


def test_camel_to_words():
    assert camel_to_words("createOrder") == "Create Order"
    assert camel_to_words("OrderController") == "Order Controller"
    assert camel_to_words("getOrderById") == "Get Order By Id"
    assert camel_to_words("getUserInfo") == "Get User Info"


def test_generate_encoding():
    assert generate_encoding("AA-01", 1) == "AA-01-001"
    assert generate_encoding("AA-01", 12) == "AA-01-012"
    assert generate_encoding("AA-02", 100) == "AA-02-100"


def test_create_excel(tmp_path):
    wb = create_workbook()
    headers = ["编号", "名称", "说明"]
    rows = [
        ["AA-01-001", "订单管理", "管理订单生命周期"],
        ["AA-01-002", "用户管理", "管理用户信息"],
    ]
    add_sheet(wb, "AA-01 应用模块清单", headers, rows)
    output = tmp_path / "test_output.xlsx"
    save_workbook(wb, output)
    assert output.exists()
    assert output.stat().st_size > 0


def test_generate_aa01(sample_project_path, tmp_path):
    project = scan_project(sample_project_path)
    project = analyze_modules(project)
    output = generate_aa01(project, tmp_path)
    assert output.exists()
    assert output.suffix == ".xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "应用域编号"
    assert ws.max_row >= 3  # header + at least 2 modules


def test_generate_aa02(sample_project_path, tmp_path):
    project = scan_project(sample_project_path)
    project = analyze_modules(project)
    endpoints = analyze_apis(project)
    output = generate_aa02(project, endpoints, tmp_path)
    assert output.exists()
    assert output.suffix == ".xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "功能项编号"
    assert ws.max_row >= 4  # header + at least 3 endpoints


def test_generate_aa03(sample_project_path, tmp_path):
    project = scan_project(sample_project_path)
    project = analyze_modules(project)
    endpoints = analyze_apis(project)
    output = generate_aa03(project, endpoints, tmp_path)
    assert output.exists()
    assert output.suffix == ".xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "功能子项编号"
    assert ws.max_row >= 2  # header + at least 1 sub-item


def test_generate_aa04(sample_project_path, tmp_path):
    project = scan_project(sample_project_path)
    project = analyze_modules(project)
    endpoints = analyze_apis(project)
    output = generate_aa04(project, endpoints, tmp_path)
    assert output.exists()
    assert output.suffix == ".xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "功能项编号"
    assert ws.cell(row=1, column=5).value == "所属微服务"
    assert ws.max_row >= 4  # header + at least 3 endpoints


def test_generate_aa05(sample_project_path, tmp_path):
    project = scan_project(sample_project_path)
    project = analyze_modules(project)
    integrations = analyze_integrations(project)
    output = generate_aa05(project, integrations, tmp_path)
    assert output.exists()
    assert output.suffix == ".xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(row=1, column=4).value == "集成类型"
    assert ws.max_row >= 2  # header + at least 1 integration


def test_generate_aa07(sample_project_path, tmp_path):
    project = scan_project(sample_project_path)
    project = analyze_modules(project)
    output = generate_aa07(project, tmp_path)
    assert output.exists()
    assert output.suffix == ".pptx"
    assert output.stat().st_size > 0


def test_generate_aa08(sample_project_path, tmp_path):
    project = scan_project(sample_project_path)
    project = analyze_modules(project)
    integrations = analyze_integrations(project)
    output = generate_aa08(project, integrations, tmp_path)
    assert output.exists()
    assert output.suffix == ".pptx"
    assert output.stat().st_size > 0


def test_llm_enhancer_disabled():
    enhancer = LLMEnhancer(enabled=False)
    result = enhancer.enhance_endpoint_name("createOrder")
    assert result == "Create Order"  # falls back to rule-based


def test_llm_enhancer_fallback_on_error():
    enhancer = LLMEnhancer(enabled=True, api_key="invalid-key")
    result = enhancer.enhance_endpoint_name("createOrder")
    assert result == "Create Order"  # graceful fallback


def test_llm_enhancer_module_description_disabled():
    enhancer = LLMEnhancer(enabled=False)
    module = Module(
        name="order",
        path=Path("/tmp/order"),
        controllers=["OrderController"],
        services=["OrderService"],
    )
    result = enhancer.enhance_module_description(module)
    assert result == "管理order相关业务"


def test_llm_enhancer_integration_description_disabled():
    enhancer = LLMEnhancer(enabled=False)
    integration = Integration(
        source_module="order",
        target_module="payment",
        integration_type=IntegrationType.HTTP,
        interface_name="PaymentClient",
        methods=["pay", "refund"],
    )
    result = enhancer.enhance_integration_description(integration)
    assert result == "order 通过 HTTP 调用 payment"
